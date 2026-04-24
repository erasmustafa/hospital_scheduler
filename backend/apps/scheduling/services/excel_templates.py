import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime, timedelta

from django.utils.timezone import localtime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TURKISH_MONTHS = {
    1: "OCAK",
    2: "SUBAT",
    3: "MART",
    4: "NISAN",
    5: "MAYIS",
    6: "HAZIRAN",
    7: "TEMMUZ",
    8: "AGUSTOS",
    9: "EYLUL",
    10: "EKIM",
    11: "KASIM",
    12: "ARALIK",
}

TURKISH_WEEKDAYS = {
    0: "Pazartesi",
    1: "Sali",
    2: "Carsamba",
    3: "Persembe",
    4: "Cuma",
    5: "Cumartesi",
    6: "Pazar",
}


def normalize_text(value):
    text = str(value or "").strip()
    text = (
        text.replace("İ", "I")
        .replace("ı", "i")
        .replace("Ş", "S")
        .replace("ş", "s")
        .replace("Ğ", "G")
        .replace("ğ", "g")
        .replace("Ü", "U")
        .replace("ü", "u")
        .replace("Ö", "O")
        .replace("ö", "o")
        .replace("Ç", "C")
        .replace("ç", "c")
    )
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _sheet_candidates(target_date):
    month_name = TURKISH_MONTHS.get(target_date.month, "")
    year = str(target_date.year)
    return [
        f"{month_name} {year}",
        f"{month_name}_{year}",
        f"{month_name}-{year}",
    ]


def _pick_target_sheet(workbook, target_date):
    candidates = [normalize_text(item) for item in _sheet_candidates(target_date)]
    for sheet in workbook.worksheets:
        normalized = normalize_text(sheet.title)
        if any(candidate in normalized for candidate in candidates):
            return sheet
    return workbook.active


def _find_header_row(worksheet):
    for row in range(1, worksheet.max_row + 1):
        cell_value = normalize_text(worksheet.cell(row=row, column=1).value)
        if "TARIH" in cell_value:
            return row
    return None


def _parse_day(value):
    if value is None:
        return None
    match = re.search(r"(\d{1,2})", str(value))
    if not match:
        return None
    day = int(match.group(1))
    return day if 1 <= day <= 31 else None


def _find_title_cell(worksheet):
    for row in range(1, min(worksheet.max_row, 15) + 1):
        for col in range(1, min(worksheet.max_column, 6) + 1):
            value = normalize_text(worksheet.cell(row=row, column=col).value)
            if "BIRIM" in value:
                return worksheet.cell(row=row, column=col)
    return None


def _assignment_staff_name(item):
    if hasattr(item, "staff_profile") and getattr(item, "staff_profile", None):
        return item.staff_profile.full_name
    if hasattr(item, "staff") and getattr(item, "staff", None):
        return item.staff.full_name
    return "-"


def _assignment_text(item):
    if hasattr(item, "start_datetime") and hasattr(item, "end_datetime"):
        start_dt = getattr(item, "start_datetime", None)
        end_dt = getattr(item, "end_datetime", None)
        if start_dt and end_dt:
            start = localtime(start_dt)
            end = localtime(end_dt)
            return f"{start.strftime('%H:%M')}-{end.strftime('%H:%M')}"

    shift_type = getattr(item, "shift_type", None)
    if shift_type and getattr(shift_type, "start_time", None) and getattr(shift_type, "end_time", None):
        return (
            f"{shift_type.start_time.strftime('%H:%M')}"
            f"-{shift_type.end_time.strftime('%H:%M')}"
        )

    if shift_type and getattr(shift_type, "name", None):
        return str(shift_type.name)

    return ""


def _shift_duration_hours(item):
    start_dt = getattr(item, "start_datetime", None)
    end_dt = getattr(item, "end_datetime", None)
    if start_dt and end_dt:
        start = localtime(start_dt)
        end = localtime(end_dt)
        return max((end - start).total_seconds() / 3600, 0)

    shift_type = getattr(item, "shift_type", None)
    if shift_type and getattr(shift_type, "start_time", None) and getattr(shift_type, "end_time", None):
        start_value = datetime.combine(item.assignment_date, shift_type.start_time)
        end_value = datetime.combine(item.assignment_date, shift_type.end_time)
        if end_value <= start_value:
            end_value += timedelta(days=1)
        return max((end_value - start_value).total_seconds() / 3600, 0)

    return 0.0


def _count_weekdays_in_range(start_date, end_date):
    day_count = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:
            day_count += 1
        current += timedelta(days=1)
    return day_count


def _get_required_hours_for_range(start_date, end_date):
    same_month_range = (
        start_date.year == end_date.year and start_date.month == end_date.month
    )
    if same_month_range:
        first_day = start_date.replace(day=1)
        if start_date.month == 12:
            next_month = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            next_month = start_date.replace(month=start_date.month + 1, day=1)
        last_day = next_month - timedelta(days=1)
        weekday_count = _count_weekdays_in_range(first_day, last_day)
    else:
        weekday_count = _count_weekdays_in_range(start_date, end_date)
    return float(weekday_count * 8)


def build_department_template_workbook(
    assignments,
    department,
    template_path,
    start_date=None,
    end_date=None,
):
    target_date = start_date or (assignments[0].assignment_date if assignments else date.today())
    workbook = load_workbook(template_path)
    worksheet = _pick_target_sheet(workbook, target_date)
    header_row = _find_header_row(worksheet)

    if not header_row:
        raise ValueError("Template does not contain a 'TARIH' header row.")

    staff_columns = {}
    for col in range(2, worksheet.max_column + 1):
        header_value = worksheet.cell(row=header_row, column=col).value
        if header_value not in (None, ""):
            staff_columns[normalize_text(header_value)] = col

    date_rows = {}
    for row in range(header_row + 1, worksheet.max_row + 1):
        day = _parse_day(worksheet.cell(row=row, column=1).value)
        if day:
            date_rows[day] = row

    for row in date_rows.values():
        for col in staff_columns.values():
            worksheet.cell(row=row, column=col).value = None

    schedule_map = defaultdict(list)
    unmatched_staff = set()

    for item in assignments:
        staff_name = _assignment_staff_name(item)
        staff_key = normalize_text(staff_name)
        target_col = staff_columns.get(staff_key)
        target_row = date_rows.get(item.assignment_date.day)

        if not target_col or not target_row:
            unmatched_staff.add(staff_name)
            continue

        schedule_map[(target_row, target_col)].append(_assignment_text(item))

    for (row, col), values in schedule_map.items():
        worksheet.cell(row=row, column=col, value="\n".join(values))

    title_cell = _find_title_cell(worksheet)
    if title_cell:
        month_name = TURKISH_MONTHS.get(target_date.month, "").upper()
        title_cell.value = (
            f"BIRIM:{department.name.upper()} "
            f"{month_name} {target_date.year} CALISMA CIZELGESI"
        )

    return workbook, worksheet, sorted(unmatched_staff)


def worksheet_to_preview(worksheet):
    merged_starts = {}
    skipped = set()
    for merged in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        merged_starts[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) != (min_row, min_col):
                    skipped.add((row, col))

    rows = []
    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for row_index in range(1, max_row + 1):
        rendered = []
        row_has_value = False
        for col_index in range(1, max_col + 1):
            if (row_index, col_index) in skipped:
                continue
            cell = worksheet.cell(row=row_index, column=col_index)
            value = "" if cell.value is None else str(cell.value)
            row_has_value = row_has_value or bool(value.strip())
            rowspan, colspan = merged_starts.get((row_index, col_index), (1, 1))

            style = []
            fill = getattr(cell.fill, "fgColor", None)
            if (
                fill
                and getattr(fill, "type", None) == "rgb"
                and fill.rgb
                and fill.rgb not in {"00000000", "FFFFFFFF", "00FFFFFF"}
            ):
                style.append(f"background:{'#' + fill.rgb[-6:]};")
            if cell.font and cell.font.bold:
                style.append("font-weight:700;")
            if cell.alignment and cell.alignment.horizontal:
                style.append(f"text-align:{cell.alignment.horizontal};")

            rendered.append(
                {
                    "value": value.replace("\n", "<br>"),
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "style": "".join(style),
                }
            )
        if row_has_value:
            rows.append(rendered)

    return {"rows": rows, "column_count": max_col, "sheet_title": worksheet.title}


def build_dynamic_assignment_workbook(
    assignments,
    department,
    start_date=None,
    end_date=None,
):
    if not assignments:
        raise ValueError("Assignments list cannot be empty.")

    actual_start = start_date or min(item.assignment_date for item in assignments)
    actual_end = end_date or max(item.assignment_date for item in assignments)
    if actual_end < actual_start:
        actual_start, actual_end = actual_end, actual_start

    staff_names = sorted(
        {
            _assignment_staff_name(item)
            for item in assignments
            if _assignment_staff_name(item) and _assignment_staff_name(item) != "-"
        }
    )
    if not staff_names:
        staff_names = ["Atanmamis Personel"]

    schedule_map = defaultdict(list)
    for item in assignments:
        schedule_map[(item.assignment_date, _assignment_staff_name(item))].append(
            _assignment_text(item)
        )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Vardiya Cizelgesi"
    worksheet.freeze_panes = "C5"
    worksheet.sheet_view.showGridLines = False

    last_column = 2 + len(staff_names)
    header_fill = PatternFill("solid", fgColor="2F5BFF")
    subheader_fill = PatternFill("solid", fgColor="EAF0FF")
    weekend_fill = PatternFill("solid", fgColor="F8FAFF")
    border_color = "D7E3FF"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    worksheet["A1"] = f"{department.name} Vardiya Cizelgesi"
    worksheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    worksheet["A1"].fill = header_fill
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    worksheet["A2"] = (
        f"Tarih Araligi: {actual_start.strftime('%d.%m.%Y')} - "
        f"{actual_end.strftime('%d.%m.%Y')}"
    )
    worksheet["A2"].font = Font(size=11, bold=True, color="24407A")
    worksheet["A2"].fill = subheader_fill
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[2].height = 22

    worksheet["A4"] = "Tarih"
    worksheet["B4"] = "Gun"
    for index, staff_name in enumerate(staff_names, start=3):
        worksheet.cell(row=4, column=index, value=staff_name)

    for col in range(1, last_column + 1):
        cell = worksheet.cell(row=4, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    worksheet.column_dimensions["A"].width = 14
    worksheet.column_dimensions["B"].width = 16
    for col in range(3, last_column + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 22

    current_date = actual_start
    row_index = 5
    while current_date <= actual_end:
        is_weekend = current_date.weekday() >= 5
        row_fill = weekend_fill if is_weekend else None

        worksheet.cell(row=row_index, column=1, value=current_date.strftime("%d.%m.%Y"))
        worksheet.cell(
            row=row_index,
            column=2,
            value=TURKISH_WEEKDAYS.get(current_date.weekday(), ""),
        )

        for col in range(1, last_column + 1):
            cell = worksheet.cell(row=row_index, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            if row_fill:
                cell.fill = row_fill

        for offset, staff_name in enumerate(staff_names, start=3):
            values = schedule_map.get((current_date, staff_name), [])
            if values:
                worksheet.cell(row=row_index, column=offset, value="\n".join(values))

        worksheet.row_dimensions[row_index].height = 34
        current_date = current_date.fromordinal(current_date.toordinal() + 1)
        row_index += 1

    summary_sheet = workbook.create_sheet("Personel Ozeti")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.freeze_panes = "A3"
    summary_sheet.merge_cells("A1:D1")
    summary_sheet["A1"] = f"{department.name} Personel Ozeti"
    summary_sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary_sheet["A1"].fill = header_fill
    summary_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary_sheet.row_dimensions[1].height = 26

    summary_headers = ["Personel", "Zorunlu Saat", "Toplam Saat", "Fazla Sure"]
    for index, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=2, column=index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 16
    summary_sheet.column_dimensions["C"].width = 16
    summary_sheet.column_dimensions["D"].width = 16

    required_hours = round(_get_required_hours_for_range(actual_start, actual_end), 1)
    staff_totals = defaultdict(float)
    for item in assignments:
        staff_totals[_assignment_staff_name(item)] += _shift_duration_hours(item)

    for row_number, staff_name in enumerate(staff_names, start=3):
        total_hours = round(staff_totals.get(staff_name, 0.0), 1)
        extra_hours = round(max(total_hours - required_hours, 0.0), 1)
        row_values = [staff_name, required_hours, total_hours, extra_hours]
        for col_number, value in enumerate(row_values, start=1):
            cell = summary_sheet.cell(row=row_number, column=col_number, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_sheet.cell(row=row_number, column=1).alignment = Alignment(
            horizontal="left", vertical="center"
        )

    return workbook, worksheet, []
